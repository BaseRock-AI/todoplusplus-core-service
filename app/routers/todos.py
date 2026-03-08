import json
import logging
from csv import DictReader
from io import BytesIO, StringIO
from pathlib import Path
from typing import Literal

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, Response, UploadFile, status
from openpyxl import Workbook, load_workbook
from pydantic import ValidationError
from sqlalchemy.orm import Session

from app.auth.dependencies import get_current_user
from app.core.config import settings
from app.db import get_db
from app.kafka_client import publisher
from app.logging_utils import Events, log_event
from app.models import AttachmentCategory, DeleteRequestStatus, User, UserRole
from app.repositories import (
    clear_todos,
    create_delete_request,
    create_todo,
    create_todo_attachment,
    create_todos_bulk,
    delete_todo,
    get_pending_delete_request,
    get_todo,
    get_todo_attachment,
    get_todo_item,
    list_todos,
    update_todo,
)
from app.schemas import (
    ToDoAttachmentOut,
    ToDoAttachmentUploadOut,
    ToDoBulkCreateResult,
    ToDoClearResponse,
    ToDoCompleteResponse,
    ToDoCreate,
    ToDoOut,
    ToDoUpdate,
)
from app.schemas_delete_request import DeleteTodoActionResponse
from app.services.storage import storage_provider

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/todos", tags=["todos"])
BULK_IMPORT_EXAMPLES_DIR = Path(__file__).resolve().parent.parent / "assets" / "bulk_import_examples"
CSV_EXAMPLE_FILENAME = "bulk-import-example.csv"
JSON_EXAMPLE_FILENAME = "bulk-import-example.json"
XLSX_EXAMPLE_FILENAME = "bulk-import-example.xlsx"


def _load_example_file_bytes(filename: str) -> bytes:
    try:
        return (BULK_IMPORT_EXAMPLES_DIR / filename).read_bytes()
    except FileNotFoundError as exc:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Missing example file: {filename}") from exc


def _build_xlsx_example_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Todos"
    sheet.append(["name", "completed"])
    sheet.append(["Write sprint summary", False])
    sheet.append(["Review backlog", True])
    sheet.append(["Plan next demo", False])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _publish_todo_created(todo_id: int, name: str, completed: bool) -> None:
    todo_payload = {"id": todo_id, "name": name, "completed": completed}
    publisher.publish(settings.topic_jira, str(todo_id), todo_payload)
    publisher.publish(settings.topic_audit, str(todo_id), {"type": "TODO", "value": json.dumps(todo_payload)})


def _parse_completed_value(raw: object) -> bool:
    if isinstance(raw, bool):
        return raw
    if raw is None or raw == "":
        return False
    if isinstance(raw, (int, float)):
        return bool(raw)
    raw_text = str(raw).strip().lower()
    if raw_text in {"true", "1", "yes", "y"}:
        return True
    if raw_text in {"false", "0", "no", "n"}:
        return False
    raise ValueError(f"Invalid completed value: {raw}")


def _parse_bulk_json_payload(payload: object) -> list[ToDoCreate]:
    items = payload.get("items") if isinstance(payload, dict) else payload
    if not isinstance(items, list):
        raise ValueError("JSON import must be a list of todo objects or {\"items\": [...]}")

    todos: list[ToDoCreate] = []
    for index, item in enumerate(items, start=1):
        try:
            todos.append(ToDoCreate.model_validate(item))
        except ValidationError as exc:
            first_error = exc.errors()[0]
            raise ValueError(f"JSON item {index}: {first_error.get('msg', 'invalid todo item')}") from exc
    return todos


def _parse_bulk_file(filename: str, data: bytes) -> list[ToDoCreate]:
    lower_name = filename.lower()
    if lower_name.endswith(".json"):
        payload = json.loads(data.decode("utf-8"))
        return _parse_bulk_json_payload(payload)

    if lower_name.endswith(".csv"):
        stream = StringIO(data.decode("utf-8-sig"))
        reader = DictReader(stream)
        if not reader.fieldnames:
            raise ValueError("CSV import requires a header row with at least a 'name' column")

        normalized_headers = {
            str(header).strip().lower(): header
            for header in reader.fieldnames
            if header is not None and str(header).strip()
        }
        if "name" not in normalized_headers:
            raise ValueError("CSV import requires a 'name' header column")

        name_header = normalized_headers["name"]
        completed_header = normalized_headers.get("completed")
        todos: list[ToDoCreate] = []
        for row_index, row in enumerate(reader, start=2):
            name_value = str(row.get(name_header, "")).strip()
            completed_raw = row.get(completed_header) if completed_header is not None else False
            completed_text = "" if completed_raw is None else str(completed_raw).strip()

            # Skip rows that are completely empty.
            if name_value == "" and completed_text == "":
                continue

            try:
                todos.append(ToDoCreate(name=name_value, completed=_parse_completed_value(completed_raw)))
            except (ValidationError, ValueError) as exc:
                raise ValueError(f"CSV row {row_index}: {exc}") from exc
        return todos

    if lower_name.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value).strip().lower() if value is not None else "" for value in rows[0]]
        header_index = {header: index for index, header in enumerate(headers)}
        if "name" not in header_index:
            raise ValueError("XLSX import requires a 'name' header column")
        completed_index = header_index.get("completed")
        todos: list[ToDoCreate] = []
        for row_index, row in enumerate(rows[1:], start=2):
            name_value = row[header_index["name"]] if header_index["name"] < len(row) else None
            completed_value = row[completed_index] if completed_index is not None and completed_index < len(row) else False
            name_text = str(name_value or "").strip()
            completed_text = "" if completed_value is None else str(completed_value).strip()

            # Skip rows that are completely empty.
            if name_text == "" and completed_text == "":
                continue

            try:
                todos.append(ToDoCreate(name=name_text, completed=_parse_completed_value(completed_value)))
            except (ValidationError, ValueError) as exc:
                raise ValueError(f"XLSX row {row_index}: {exc}") from exc
        return todos

    raise ValueError("Unsupported file type. Use .json, .csv, or .xlsx")


@router.get("", response_model=list[ToDoOut])
def get_todos(
    scope: Literal["all", "done", "pending"] = Query(default="all"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> list[ToDoOut]:
    completed_filter = None
    if scope == "done":
        completed_filter = True
    elif scope == "pending":
        completed_filter = False
    return list_todos(db, completed_filter=completed_filter)


@router.delete("/clear", response_model=ToDoClearResponse)
def clear_todo_items(
    scope: Literal["all", "done", "pending"] = Query(default="all"),
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> ToDoClearResponse:
    deleted_count = clear_todos(db, scope=scope)
    return ToDoClearResponse(scope=scope, deleted_count=deleted_count)


@router.get("/{todo_id}", response_model=ToDoOut)
def get_todo_by_id(todo_id: int, _: User = Depends(get_current_user), db: Session = Depends(get_db)) -> ToDoOut:
    todo = get_todo(db, todo_id)
    if not todo:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"ToDo item {todo_id} not found")
    return todo


@router.post("", response_model=ToDoOut, status_code=status.HTTP_201_CREATED)
def create_todo_item(payload: ToDoCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> ToDoOut:
    log_event(
        logger,
        logging.INFO,
        Events.TODO_CREATE_REQUEST,
        name=payload.name,
        completed=payload.completed,
    )

    todo = create_todo(db, payload, current_user.id)

    log_event(logger, logging.INFO, Events.TODO_CREATED_DB, todo_id=todo.id, name=todo.name, completed=todo.completed)
    _publish_todo_created(todo.id, todo.name, todo.completed)
    return get_todo(db, todo.id)


@router.get("/bulk-import/examples/tabular")
def download_bulk_import_tabular_example(
    format: Literal["csv", "xlsx"] = "csv",
    _: User = Depends(get_current_user),
) -> Response:
    if format == "csv":
        body = _load_example_file_bytes(CSV_EXAMPLE_FILENAME)
        headers = {"Content-Disposition": f'attachment; filename="{CSV_EXAMPLE_FILENAME}"'}
        return Response(content=body, media_type="text/csv", headers=headers)

    body = _build_xlsx_example_bytes()
    headers = {"Content-Disposition": f'attachment; filename="{XLSX_EXAMPLE_FILENAME}"'}
    return Response(
        content=body,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/bulk-import/examples/json")
def download_bulk_import_json_example(_: User = Depends(get_current_user)) -> Response:
    body = _load_example_file_bytes(JSON_EXAMPLE_FILENAME)
    headers = {"Content-Disposition": f'attachment; filename="{JSON_EXAMPLE_FILENAME}"'}
    return Response(content=body, media_type="application/json", headers=headers)


@router.post("/bulk-import", response_model=ToDoBulkCreateResult, status_code=status.HTTP_201_CREATED)
async def bulk_import_todos(
    request: Request,
    file: UploadFile | None = File(default=None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> ToDoBulkCreateResult:
    try:
        if file is not None:
            data = await file.read()
            if len(data) > settings.upload_max_bytes:
                raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="Upload exceeds maximum allowed size")
            todos = _parse_bulk_file(file.filename or "", data)
        else:
            raw_body = await request.body()
            if len(raw_body) > settings.upload_max_bytes:
                raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="Upload exceeds maximum allowed size")
            payload = json.loads(raw_body.decode("utf-8"))
            todos = _parse_bulk_json_payload(payload)
    except (ValueError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=str(exc)) from exc

    if not todos:
        return ToDoBulkCreateResult(created_count=0, ids=[])

    created = create_todos_bulk(db, todos, current_user.id)
    for todo in created:
        _publish_todo_created(todo.id, todo.name, todo.completed)
    return ToDoBulkCreateResult(created_count=len(created), ids=[todo.id for todo in created])


@router.post("/{todo_id}/complete", response_model=ToDoCompleteResponse)
def complete_todo_item(
    todo_id: int,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> ToDoCompleteResponse:
    todo = get_todo_item(db, todo_id)
    if not todo:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"ToDo item {todo_id} not found")

    updated = update_todo(db, todo, ToDoUpdate(completed=True))
    return ToDoCompleteResponse(id=updated.id, completed=updated.completed)


@router.post("/{todo_id}/attachments/upload", response_model=ToDoAttachmentUploadOut, status_code=status.HTTP_201_CREATED)
@router.post(
    "/{todo_id}/attachments",
    response_model=ToDoAttachmentUploadOut,
    status_code=status.HTTP_201_CREATED,
    include_in_schema=False,
)
async def upload_todo_attachment(
    todo_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> ToDoAttachmentUploadOut:
    todo = get_todo_item(db, todo_id)
    if not todo:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"ToDo item {todo_id} not found")

    data = await file.read()
    if len(data) > settings.upload_max_bytes:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="Upload exceeds maximum allowed size")

    stored_file = storage_provider.save_bytes(
        namespace=f"todos/{todo_id}/attachments",
        filename=file.filename or "upload.bin",
        data=data,
        content_type=file.content_type,
    )
    attachment = create_todo_attachment(
        db=db,
        todo_id=todo_id,
        category=AttachmentCategory.COMPLETION_PROOF,
        storage_key=stored_file.storage_key,
        original_filename=stored_file.original_filename,
        content_type=stored_file.content_type,
        size_bytes=stored_file.size_bytes,
        uploaded_by_user_id=current_user.id,
    )
    return ToDoAttachmentUploadOut(
        attachment_id=attachment.id,
        todo_id=attachment.todo_id,
        filename=attachment.original_filename,
        content_type=attachment.content_type,
        size_bytes=attachment.size_bytes,
    )


@router.get("/{todo_id}/attachments/{attachment_id}", response_model=ToDoAttachmentOut)
def get_todo_attachment_details(
    todo_id: int,
    attachment_id: int,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> ToDoAttachmentOut:
    attachment = get_todo_attachment(db, attachment_id)
    if attachment is None or attachment.todo_id != todo_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Attachment not found")
    return ToDoAttachmentOut(
        id=attachment.id,
        todo_id=attachment.todo_id,
        filename=attachment.original_filename,
        content_type=attachment.content_type,
        size_bytes=attachment.size_bytes,
    )


@router.get("/{todo_id}/attachments/{attachment_id}/download")
def download_todo_attachment(
    todo_id: int,
    attachment_id: int,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Response:
    attachment = get_todo_attachment(db, attachment_id)
    if attachment is None or attachment.todo_id != todo_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Attachment not found")
    if not storage_provider.exists(attachment.storage_key):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Attachment file missing in storage")

    body = storage_provider.read_bytes(attachment.storage_key)
    headers = {"Content-Disposition": f'attachment; filename="{attachment.original_filename}"'}
    return Response(content=body, media_type=attachment.content_type or "application/octet-stream", headers=headers)


@router.put("/{todo_id}", response_model=ToDoOut)
def update_todo_item(
    todo_id: int,
    payload: ToDoUpdate,
    _: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> ToDoOut:
    todo = get_todo_item(db, todo_id)
    if not todo:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"ToDo item {todo_id} not found")
    updated_todo = update_todo(db, todo, payload)
    return get_todo(db, updated_todo.id)


@router.delete("/{todo_id}", response_model=DeleteTodoActionResponse)
def delete_todo_item(
    todo_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> DeleteTodoActionResponse:
    todo = get_todo_item(db, todo_id)
    if not todo:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"ToDo item {todo_id} not found")

    if current_user.role == UserRole.ADMIN:
        delete_todo(db, todo)
        return DeleteTodoActionResponse(action="deleted", message=f"ToDo item {todo_id} deleted")

    existing_request = get_pending_delete_request(db, todo_id, current_user.id)
    if existing_request:
        return DeleteTodoActionResponse(
            action="pending_approval",
            message="Delete request already pending admin approval",
            delete_request_id=existing_request.id,
        )

    delete_request = create_delete_request(db, todo_id=todo_id, requested_by_user_id=current_user.id)
    return DeleteTodoActionResponse(
        action=DeleteRequestStatus.PENDING,
        message="Delete request submitted for admin approval",
        delete_request_id=delete_request.id,
    )
